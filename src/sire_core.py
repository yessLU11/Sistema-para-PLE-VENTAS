# src/sire_core.py
"""
CONVERTIDOR SIRE/PLE - SUNAT
Versión 2.1 - Motor de Procesamiento Core (Doble Formato + Archivos Grandes)
Área de Tributación - Banco de la Nación del Perú
"""

import os
import logging
import gc
from datetime import datetime
from pathlib import Path
from io import StringIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Asegurar directorios mínimos para el correcto funcionamiento del core
Path("input_files").mkdir(exist_ok=True)
Path("output_files").mkdir(exist_ok=True)
Path("logs").mkdir(exist_ok=True)

# ============================================================================
# SISTEMA DE LOGGING
# ============================================================================

class ConversionLogger:
    """Logger especializado para conversiones en background"""
    
    def __init__(self):
        self.logger = self._setup_logger()
        self.stats = {}
    
    def _setup_logger(self):
        logger = logging.getLogger('SIREConverter')
        logger.setLevel(logging.DEBUG)
        
        # Limpiar handlers existentes para evitar duplicados en Streamlit
        if logger.handlers:
            logger.handlers.clear()
        
        formatter = logging.Formatter(
            '[%(asctime)s] %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Console handler (Terminal)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        
        # File handler (Logs locales)
        log_file = f"logs/conversion_{datetime.now().strftime('%Y%m%d')}.log"
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        return logger
    
    def info(self, msg):
        self.logger.info(msg)
    
    def error(self, msg):
        self.logger.error(msg)
    
    def warning(self, msg):
        self.logger.warning(msg)

# Inicializar logger global del Core
conv_logger = ConversionLogger()

# ============================================================================
# VALIDADORES
# ============================================================================

class SIREValidator:
    """Validador de datos SIRE"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def validate_columns(self, df):
        """Validar columnas críticas"""
        critical_cols = ['Ruc', 'Razón Social', 'Total CP', 'Moneda']
        missing = [col for col in critical_cols if col in df.columns]
        
        # Si tiene alguna de las columnas críticas, está bien
        if not missing:
            return True
        
        # Warning pero no error - puede ser formato sin headers
        if len(df.columns) >= 20:
            self.warnings.append("Archivo sin encabezados detectados - usando formato automático")
            return True
            
        self.errors.append(f"Columnas faltantes: {', '.join(missing)}")
        return True  # No bloqueamos, solo advertimos
    
    def validate_row_count(self, df):
        """Validar que hay datos"""
        if len(df) == 0:
            self.errors.append("El archivo no contiene datos")
            return False
        return True
    
    def validate_all(self, df):
        """Ejecutar todas las validaciones"""
        self.errors = []
        self.warnings = []
        
        self.validate_columns(df)
        self.validate_row_count(df)
        
        return {
            'is_valid': len(df) > 0,
            'errors': self.errors,
            'warnings': self.warnings,
            'total_rows': len(df),
            'total_columns': len(df.columns)
        }

# ============================================================================
# PROCESADOR TXT - FORMATO 1 (CON ENCABEZADOS)
# ============================================================================

class TXTProcessorConEncabezado:
    """Procesador para archivos TXT SIRE CON encabezados"""
    
    def __init__(self, max_size_mb=500):
        self.max_size_mb = max_size_mb
        self.file_size = None
        self.row_count = None
    
    def validate_file_size(self, file_path):
        """Validar tamaño del archivo"""
        file_size = os.path.getsize(file_path)
        self.file_size = file_size / (1024 * 1024)
        
        if self.file_size > self.max_size_mb:
            msg = f"Archivo muy grande: {self.file_size:.2f}MB (máx: {self.max_size_mb}MB)"
            conv_logger.warning(msg)
            return False, msg
        
        return True, None
    
    def read_txt(self, file_path):
        """Leer archivo TXT con encabezados"""
        try:
            is_valid, size_error = self.validate_file_size(file_path)
            if not is_valid:
                return None, size_error
            
            conv_logger.info(f"Iniciando lectura CON encabezados: {file_path} ({self.file_size:.2f}MB)")
            
            # Siempre usar chunks para archivos > 10 MB
            chunks = []
            chunk_size = 50000
            
            for i, chunk in enumerate(pd.read_csv(
                file_path,
                sep='|',
                encoding='latin-1',
                decimal='.',
                skipinitialspace=True,
                engine='python',
                on_bad_lines='skip',
                chunksize=chunk_size
            )):
                chunks.append(chunk)
                filas = sum(len(c) for c in chunks)
                if (i + 1) % 5 == 0:
                    conv_logger.info(f"Procesado chunk {i+1}: {filas:,} filas...")
            
            if not chunks:
                return None, "No se pudieron leer datos"
            
            df = pd.concat(chunks, ignore_index=True)
            
            # Limpiar columnas
            df.columns = df.columns.str.strip()
            if len(df.columns) > 0 and df.columns[-1] in ['', 'Unnamed: 0']:
                df = df.iloc[:, :-1]
            
            self.row_count = len(df)
            conv_logger.info(f"Total leído: {self.row_count:,} filas")
            
            # Limpiar memoria
            del chunks
            gc.collect()
            
            return df, None
            
        except Exception as e:
            error_msg = f"Error al procesar: {str(e)}"
            conv_logger.error(error_msg)
            return None, error_msg

# ============================================================================
# PROCESADOR TXT - FORMATO 2 (SIN ENCABEZADOS)
# ============================================================================

class TXTProcessorSinEncabezado:
    """Procesador para archivos TXT SIRE SIN encabezados (con LOG inicial)"""
    
    # 36 columnas basadas en el formato real del archivo (no 31)
    COLUMNAS = [
        'Tipo', 'Periodo', 'IDComprobante', 'Serie', 'FechaEmision',
        'FechaVencimiento', 'TipoDoc', 'CodigoEstablecimiento', 'NumeroCorrelativo',
        'campo_10', 'TipoDocCliente', 'NumeroDocCliente', 'RazonSocialCliente',
        'MontoOperacionesExoneradas', 'MontoOperacionesGravadas', 'MontoIGV',
        'CodigoCentroCosto', 'MontoIGVRetenido', 'MontoReversiones',
        'MontoOtrosConceptos', 'MontoPercepciones', 'MontoDetraccion',
        'MontoTotal', 'campo_23', 'campo_24', 'campo_25',
        'campo_26', 'campo_27', 'campo_28', 'campo_29',
        'campo_30', 'campo_31', 'campo_32', 'campo_33',
        'campo_34', 'campo_35', 'campo_36'
    ]
    
    def __init__(self, max_size_mb=500):
        self.max_size_mb = max_size_mb
        self.file_size = None
        self.row_count = None
    
    def validate_file_size(self, file_path):
        """Validar tamaño del archivo"""
        file_size = os.path.getsize(file_path)
        self.file_size = file_size / (1024 * 1024)
        
        if self.file_size > self.max_size_mb:
            msg = f"Archivo muy grande: {self.file_size:.2f}MB (máx: {self.max_size_mb}MB)"
            conv_logger.warning(msg)
            return False, msg
        
        return True, None
    
    def read_txt(self, file_path):
        """Leer archivo TXT sin encabezados - VERSION CORREGIDA"""
        try:
            is_valid, size_error = self.validate_file_size(file_path)
            if not is_valid:
                return None, size_error
            
            conv_logger.info(f"Iniciando lectura SIN encabezados: {file_path} ({self.file_size:.2f}MB)")
            
            with open(file_path, 'r', encoding='latin-1', errors='replace') as f:
                todas_lineas = f.readlines()
            
            conv_logger.info(f"Total líneas en archivo: {len(todas_lineas):,}")
            
            # Buscar la primera línea de datos REALES
            linea_inicio = 0
            for i, linea in enumerate(todas_lineas):
                linea_limpia = linea.strip()
                
                # Es datos si empieza con número y contiene suficientes delimitadores
                if (linea_limpia and 
                    linea_limpia[0].isdigit() and 
                    linea_limpia.count('|') >= 20):
                    linea_inicio = i
                    conv_logger.info(f"🔥 Datos encontrados en línea {i+1}")
                    conv_logger.info(f"    Primera línea de datos: {linea_limpia[:80]}...")
                    break
            
            if linea_inicio == 0:
                conv_logger.warning("⚠️ No se encontró delimitador claro, interpretando todo como datos")
            
            lineas_datos = todas_lineas[linea_inicio:]
            conv_logger.info(f"📊 Líneas de datos a procesar: {len(lineas_datos):,}")
            
            contenido = ''.join(lineas_datos)
            
            df = pd.read_csv(
                StringIO(contenido),
                sep='|',
                encoding='latin-1',
                header=None,
                names=self.COLUMNAS,
                dtype=str,  # Forzar string evita rupturas por pérdida de ceros a la izquierda
                on_bad_lines='skip'
            )
            
            # Eliminar última columna si está vacía (por el pipe '|' de cierre)
            if len(df.columns) > 0:
                ultima_col = df.columns[-1]
                if df[ultima_col].isna().all() or (df[ultima_col].astype(str).str.strip() == '').all():
                    df = df.iloc[:, :-1]
                    conv_logger.info("✅ Columna vacía final eliminada")
            
            self.row_count = len(df)
            conv_logger.info(f"✅ Total leído: {self.row_count:,} filas")
            
            # Liberar memoria intermedia
            del todas_lineas, lineas_datos
            gc.collect()
            
            return df, None
            
        except MemoryError:
            conv_logger.error("Memoria insuficiente para procesar el archivo")
            return None, "Memoria insuficiente. Archivo muy grande."
        except Exception as e:
            error_msg = f"Error al procesar: {str(e)}"
            conv_logger.error(error_msg)
            return None, error_msg

# ============================================================================
# GENERADOR EXCEL
# ============================================================================

class ExcelGenerator:
    """Generador Excel con múltiples hojas"""
    
    def __init__(self, max_rows=1_000_000):
        self.max_rows = max_rows
        self.sheets_created = 0
    
    def split_dataframe(self, df):
        """Dividir en chunks si excede máximo de filas por hoja"""
        if len(df) <= self.max_rows:
            return [df]
        
        chunks = []
        num_chunks = (len(df) // self.max_rows) + 1
        
        for i in range(num_chunks):
            start_idx = i * self.max_rows
            end_idx = min((i + 1) * self.max_rows, len(df))
            chunks.append(df.iloc[start_idx:end_idx].reset_index(drop=True))
        
        conv_logger.info(f"Dividido en {len(chunks)} hojas de hasta {self.max_rows:,} filas")
        return chunks
    
    def format_header(self, ws):
        """Formatear encabezados"""
        fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            if cell.value:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
    
    def auto_fit_columns(self, ws):
        """Ajustar ancho de columnas automáticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Muestrear las primeras 100 filas para agilizar el rendimiento
            for cell in list(column)[:100]:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_excel(self, df, output_path):
        """Crear Excel con múltiples hojas si es necesario"""
        try:
            df_chunks = self.split_dataframe(df)
            self.sheets_created = len(df_chunks)
            
            conv_logger.info(f"Generando Excel con {len(df_chunks)} hoja(s)...")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Construcción de hojas de datos
            for idx, chunk in enumerate(df_chunks, 1):
                sheet_name = f"Datos_{idx}" if len(df_chunks) > 1 else "Datos"
                ws = wb.create_sheet(sheet_name)
                
                # Inserción veloz por filas
                for r_idx, row in enumerate(dataframe_to_rows(chunk, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                self.format_header(ws)
                self.auto_fit_columns(ws)
                conv_logger.info(f"Hoja {idx} completada")
            
            # Construcción de hoja resumen
            summary_ws = wb.create_sheet("Resumen", 0)
            self._create_summary(summary_ws, df_chunks, df)
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            
            msg = f"Excel creado: {self.sheets_created} hoja(s)"
            conv_logger.info(msg)
            return True, msg
            
        except Exception as e:
            error_msg = f"Error al generar Excel: {str(e)}"
            conv_logger.error(error_msg)
            return False, error_msg
    
    def _create_summary(self, ws, chunks, df):
        """Crear la carátula resumen del Excel"""
        ws['A1'] = "RESUMEN DE CONVERSIÓN"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        data = [
            ['Total de Registros', len(df)],
            ['Total de Columnas', len(df.columns)],
            ['Número de Hojas', len(chunks)],
            ['Máx. Filas/Hoja', f"{self.max_rows:,}"]
        ]
        
        # Evaluar sumatoria dinámica si existe columna de dinero/totales
        columnas_total = ['MontoTotal', 'Total CP', 'Total', 'Monto Operacion Gravadas', 'MontoIGV']
        for col_name in columnas_total:
            if col_name in df.columns:
                try:
                    total = pd.to_numeric(df[col_name], errors='coerce').sum()
                    data.append(['Monto Total', f"{total:,.2f}"])
                except:
                    pass
                break
        
        for row_idx, (label, value) in enumerate(data, 3):
            ws[f'A{row_idx}'] = label
            ws[f'B{row_idx}'] = value
            ws[f'A{row_idx}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20